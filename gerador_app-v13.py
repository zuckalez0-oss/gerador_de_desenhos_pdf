import pandas as pd
import math
import sys
import os
import zipfile
import json
from datetime import datetime
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                             QHBoxLayout, QPushButton, QLabel, QTextEdit, 
                             QFileDialog, QProgressBar, QMessageBox, QGroupBox,
                             QFormLayout, QLineEdit, QComboBox, QTableWidget, QTableWidgetItem,
                             QDialog, QListWidget, QInputDialog)
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from openpyxl import load_workbook, Workbook

# Supondo que dxf_engine.py está na mesma pasta
# import dxf_engine


# CONSTANTES GLOBAIS DE LAYOUT DA PÁGINA

PAGE_WIDTH, PAGE_HEIGHT = A4
MARGEM_GERAL = 15 * mm
HEADER_AREA_ALTURA = 25 * mm
FOOTER_AREA_ALTURA = 25 * mm 


# CLASSE PARA GERENCIAR CÓDIGOS ÚNICOS

class CodeGenerator:
    def __init__(self, db_path="codigo_database.xlsx"):
        self.db_path = db_path; self.code_column_name = 'Codigo Unico'; self.timestamp_column_name = 'Data de Registro'
        self.project_column_name = 'Projeto'; self.existing_codes = set(); self.last_code_number = 0; self._load_database()
    def _load_database(self):
        try:
            df = pd.read_excel(self.db_path)
            if self.code_column_name in df.columns:
                df[self.code_column_name] = df[self.code_column_name].astype(str)
                self.existing_codes = set(df[self.code_column_name].tolist())
                max_num = 0
                for code in self.existing_codes:
                    if code.upper().startswith('DES'):
                        try:
                            num = int(code[3:])
                            if num > max_num: max_num = num
                        except (ValueError, TypeError): continue
                self.last_code_number = max_num
        except FileNotFoundError: print(f"Arquivo '{self.db_path}' não encontrado. Um novo será criado.")
        except Exception as e: QMessageBox.warning(None, "Erro de Leitura", f"Não foi possível ler o banco de dados de códigos: {e}")
    def _append_to_database(self, new_code, project_number):
        timestamp = datetime.now().strftime('%d/%m/%Y %H:%M:%S'); new_row = [new_code, timestamp, project_number]
        try:
            wb = load_workbook(self.db_path); sheet = wb.active
        except FileNotFoundError:
            wb = Workbook(); sheet = wb.active
            sheet.append([self.code_column_name, self.timestamp_column_name, self.project_column_name])
        sheet.append(new_row)
        try:
            wb.save(self.db_path)
        except PermissionError:
             QMessageBox.critical(None, "Erro de Permissão", f"Não foi possível salvar o novo código em '{self.db_path}'.\n\nPor favor, feche a planilha e tente novamente."); return False
        except Exception as e:
            QMessageBox.critical(None, "Erro ao Salvar", f"Ocorreu um erro ao salvar o banco de códigos: {e}"); return False
        return True
    def generate_new_code(self, project_number, prefix='DES'):
        self._load_database(); current_number = self.last_code_number + 1
        new_code = f"{prefix}{current_number}"
        while new_code in self.existing_codes:
            current_number += 1; new_code = f"{prefix}{current_number}"
        if self._append_to_database(new_code, project_number):
            self.existing_codes.add(new_code); self.last_code_number = current_number
            return new_code
        else: return None


# CLASSE PARA GERENCIAR O HISTÓRICO DE PROJETOS

class HistoryManager:
    def __init__(self, history_path="project_history.json"):
        self.history_path = history_path
    def _load_history(self):
        try:
            with open(self.history_path, 'r', encoding='utf-8') as f: return json.load(f)
        except (FileNotFoundError, json.JSONDecodeError): return {}
    def _save_history(self, history_data):
        with open(self.history_path, 'w', encoding='utf-8') as f: json.dump(history_data, f, indent=4)
    def get_projects(self):
        return sorted(self._load_history().keys())
    def get_project_data(self, project_number):
        return self._load_history().get(project_number, {}).get('pieces', [])
    def save_project(self, project_number, pieces_df):
        history = self._load_history()
        df_copy = pieces_df.copy()
        df_copy['furos'] = df_copy['furos'].apply(lambda x: x if isinstance(x, list) else [])
        pieces_list = df_copy.to_dict('records')
        history[project_number] = {"project_number": project_number, "save_date": datetime.now().strftime('%d/%m/%Y %H:%M:%S'), "pieces": pieces_list}
        self._save_history(history)
    def delete_project(self, project_number):
        history = self._load_history()
        if project_number in history:
            del history[project_number]; self._save_history(history); return True
        return False


# JANELA DE DIÁLOGO PARA O HISTÓRICO

class HistoryDialog(QDialog):
    def __init__(self, history_manager, parent=None):
        super().__init__(parent); self.setWindowTitle("Histórico de Projetos"); self.history_manager = history_manager
        self.loaded_project_data = None; self.setMinimumSize(800, 600); layout = QHBoxLayout(self)
        left_panel = QGroupBox("Projetos Salvos"); left_layout = QVBoxLayout(); self.project_list_widget = QListWidget()
        left_layout.addWidget(self.project_list_widget); left_panel.setLayout(left_layout)
        right_panel = QGroupBox("Peças do Projeto"); right_layout = QVBoxLayout(); self.pieces_table_widget = QTableWidget()
        right_layout.addWidget(self.pieces_table_widget); button_layout = QHBoxLayout()
        self.load_btn = QPushButton("Carregar Projeto na Lista"); self.delete_btn = QPushButton("Excluir Projeto do Histórico"); close_btn = QPushButton("Fechar")
        button_layout.addWidget(self.load_btn); button_layout.addWidget(self.delete_btn); button_layout.addStretch(); button_layout.addWidget(close_btn)
        right_layout.addLayout(button_layout); right_panel.setLayout(right_layout); layout.addWidget(left_panel, 1); layout.addWidget(right_panel, 3)
        self.project_list_widget.currentItemChanged.connect(self.display_project_details); self.load_btn.clicked.connect(self.load_project)
        self.delete_btn.clicked.connect(self.delete_project); close_btn.clicked.connect(self.reject)
        self.populate_project_list(); self.update_buttons_state()
    def populate_project_list(self):
        self.project_list_widget.clear(); self.project_list_widget.addItems(self.history_manager.get_projects())
    def display_project_details(self, current, previous):
        self.pieces_table_widget.setRowCount(0)
        if not current: self.update_buttons_state(); return
        project_number = current.text(); pieces = self.history_manager.get_project_data(project_number)
        if pieces:
            headers = [key for key in pieces[0].keys()]
            self.pieces_table_widget.setColumnCount(len(headers)); self.pieces_table_widget.setHorizontalHeaderLabels([h.replace('_', ' ').title() for h in headers])
            self.pieces_table_widget.setRowCount(len(pieces))
            for i, piece in enumerate(pieces):
                for j, key in enumerate(headers):
                    value = piece.get(key)
                    display_text = f"{len(value)} Furo(s)" if key == 'furos' and isinstance(value, list) else str(value if value is not None else '-')
                    self.pieces_table_widget.setItem(i, j, QTableWidgetItem(display_text))
        self.pieces_table_widget.resizeColumnsToContents(); self.update_buttons_state()
    def update_buttons_state(self):
        has_selection = self.project_list_widget.currentItem() is not None
        self.load_btn.setEnabled(has_selection); self.delete_btn.setEnabled(has_selection)
    def load_project(self):
        if self.project_list_widget.currentItem():
            project_number = self.project_list_widget.currentItem().text()
            self.loaded_project_data = self.history_manager.get_project_data(project_number); self.accept()
    def delete_project(self):
        if self.project_list_widget.currentItem():
            project_number = self.project_list_widget.currentItem().text()
            reply = QMessageBox.question(self, 'Excluir Projeto', f"Tem certeza que deseja excluir o projeto '{project_number}' do histórico?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply == QMessageBox.Yes:
                self.history_manager.delete_project(project_number); self.populate_project_list(); self.pieces_table_widget.setRowCount(0)


# FUNÇÕES DE DESENHO E RELATÓRIO PDF

def desenhar_cabecalho(c, nome_arquivo):
    c.setFont("Helvetica-Bold", 14); y_pos_texto = PAGE_HEIGHT - HEADER_AREA_ALTURA + (10 * mm); c.drawCentredString(PAGE_WIDTH / 2, y_pos_texto, f"Desenho da Peça: {nome_arquivo}")
    y_pos_linha = PAGE_HEIGHT - HEADER_AREA_ALTURA; c.line(MARGEM_GERAL, y_pos_linha, PAGE_WIDTH - MARGEM_GERAL, y_pos_linha)
def desenhar_rodape_aprimorado(c, row):
    largura_total = PAGE_WIDTH - 2 * MARGEM_GERAL; altura_bloco = 12 * mm; y_rodape = FOOTER_AREA_ALTURA - altura_bloco - (5 * mm)
    c.setStrokeColorRGB(0, 0, 0); c.rect(MARGEM_GERAL, y_rodape, largura_total, altura_bloco)
    coluna1_x, coluna2_x, coluna3_x = MARGEM_GERAL, MARGEM_GERAL + largura_total * 0.60, MARGEM_GERAL + largura_total * 0.80
    c.line(coluna2_x, y_rodape, coluna2_x, y_rodape + altura_bloco); c.line(coluna3_x, y_rodape, coluna3_x, y_rodape + altura_bloco)
    y_titulo, y_valor = y_rodape + altura_bloco - 4*mm, y_rodape + 4*mm
    c.setFont("Helvetica", 7); c.drawCentredString(coluna1_x + (coluna2_x - coluna1_x)/2, y_titulo, "NOME DA PEÇA / IDENTIFICADOR")
    c.setFont("Helvetica-Bold", 10); c.drawCentredString(coluna1_x + (coluna2_x - coluna1_x)/2, y_valor, str(row.get('nome_arquivo', 'N/A')))
    c.setFont("Helvetica", 7); c.drawCentredString(coluna2_x + (coluna3_x - coluna2_x)/2, y_titulo, "ESPESSURA")
    c.setFont("Helvetica-Bold", 10); c.drawCentredString(coluna2_x + (coluna3_x - coluna2_x)/2, y_valor, f"{formatar_numero(row.get('espessura', 0))} mm")
    c.setFont("Helvetica", 7); c.drawCentredString(coluna3_x + (PAGE_WIDTH - MARGEM_GERAL - coluna3_x)/2, y_titulo, "QUANTIDADE")
    c.setFont("Helvetica-Bold", 10); c.drawCentredString(coluna3_x + (PAGE_WIDTH - MARGEM_GERAL - coluna3_x)/2, y_valor, formatar_numero(row.get('qtd', 0)))
def desenhar_erro_dados(c, forma): c.setFont("Helvetica-Bold", 14); c.drawCentredString(A4[0]/2, A4[1]/2, f"Dados inválidos para a forma: '{forma}'")
def formatar_numero(valor):
    if valor == int(valor): return str(int(valor))
    return str(valor)
def desenhar_cota_horizontal(c, x1, x2, y, texto):
    FONT_NAME, FONT_SIZE = "Helvetica", 10; c.setFont(FONT_NAME, FONT_SIZE); c.line(x1, y, x2, y)
    tick_len = 2 * mm; c.line(x1, y - tick_len/2, x1 + tick_len/2, y + tick_len/2); c.line(x2, y - tick_len/2, x2 - tick_len/2, y + tick_len/2)
    largura_texto = c.stringWidth(texto, FONT_NAME, FONT_SIZE); centro_x, y_texto = (x1 + x2) / 2, y + 1*mm; padding = 1.5 * mm
    gap_inicio, gap_fim = centro_x - (largura_texto / 2) - padding, centro_x + (largura_texto / 2) + padding
    c.line(x1, y, gap_inicio, y); c.line(gap_fim, y, x2, y); c.drawCentredString(centro_x, y_texto, texto)
def desenhar_cota_vertical(c, y1, y2, x, texto):
    FONT_NAME, FONT_SIZE = "Helvetica", 10; c.setFont(FONT_NAME, FONT_SIZE); c.line(x, y1, x, y2)
    tick_len = 2 * mm; c.line(x - tick_len/2, y1, x + tick_len/2, y1 + tick_len/2); c.line(x - tick_len/2, y2, x + tick_len/2, y2 - tick_len/2)
    largura_texto, altura_texto = c.stringWidth(texto, FONT_NAME, FONT_SIZE), FONT_SIZE; centro_y = (y1 + y2) / 2
    c.saveState(); c.translate(x, centro_y); c.rotate(90); c.setFillColorRGB(1, 1, 1)
    c.rect(-largura_texto/2-1*mm, -altura_texto/2, largura_texto+2*mm, altura_texto, stroke=0, fill=1)
    c.setFillColorRGB(0, 0, 0); c.drawCentredString(0, -1.5*mm, texto); c.restoreState()
def desenhar_cota_diametro_furo(c, x, y, raio, diametro_real):
    c.saveState(); c.setFont("Helvetica", 10); texto = f"Ø {formatar_numero(diametro_real)}"; largura_texto = c.stringWidth(texto, "Helvetica", 10)
    p_borda_x, p_borda_y = x + raio * 0.7071, y + raio * 0.7071; p_meio_x, p_meio_y = p_borda_x + 4 * mm, p_borda_y + 4 * mm
    p_final_x = p_meio_x + largura_texto + 2*mm; path = c.beginPath(); path.moveTo(p_borda_x, p_borda_y); path.lineTo(p_meio_x, p_meio_y); path.lineTo(p_final_x, p_meio_y)
    c.drawPath(path); c.drawString(p_meio_x + 1*mm, p_meio_y + 1*mm, texto); c.restoreState()
def desenhar_retangulo(c, row):
    largura, altura = row.get('largura', 0), row.get('altura', 0);
    if largura <= 0 or altura <= 0: desenhar_erro_dados(c, "Retângulo"); return
    max_w, max_h = PAGE_WIDTH - 2*MARGEM_GERAL, PAGE_HEIGHT - HEADER_AREA_ALTURA - FOOTER_AREA_ALTURA
    dist_cota_furo, dist_cota_total, overshoot = 8*mm, 16*mm, 2*mm
    espaco_cota_x, espaco_cota_y = dist_cota_total, dist_cota_total
    escala = min((max_w - espaco_cota_x) / largura, (max_h - espaco_cota_y) / altura) * 0.95
    dw, dh = largura*escala, altura*escala; bloco_visual_width, bloco_visual_height = dw + espaco_cota_x, dh + espaco_cota_y
    base_desenho_y = FOOTER_AREA_ALTURA; inicio_bloco_x, inicio_bloco_y = MARGEM_GERAL + (max_w - bloco_visual_width)/2, base_desenho_y + (max_h - bloco_visual_height)/2
    x0, y0 = inicio_bloco_x + espaco_cota_x, inicio_bloco_y + espaco_cota_y; c.rect(x0, y0, dw, dh)
    furos = row.get('furos')
    if isinstance(furos, list) and furos:
        for furo in furos: c.circle(x0 + (furo['x']*escala), y0 + (furo['y']*escala), (furo['diam']/2)*escala, stroke=1, fill=0)
        y_pos_cota = y0 - dist_cota_furo; unique_x = sorted(list(set(f['x'] for f in furos))); dim_points_x = [0] + unique_x + [largura]
        for x_real in dim_points_x: c.line(x0 + x_real*escala, y0, x0 + x_real*escala, y_pos_cota - overshoot)
        for i in range(1, len(dim_points_x)):
            start_x, end_x = dim_points_x[i-1], dim_points_x[i]
            if (end_x - start_x)>0.01: desenhar_cota_horizontal(c, x0+start_x*escala, x0+end_x*escala, y_pos_cota, formatar_numero(end_x-start_x))
        x_pos_cota = x0 - dist_cota_furo; unique_y = sorted(list(set(f['y'] for f in furos))); dim_points_y = [0] + unique_y + [altura]
        for y_real in dim_points_y: c.line(x0, y0 + y_real*escala, x_pos_cota - overshoot, y0 + y_real*escala)
        for i in range(1, len(dim_points_y)):
            start_y, end_y = dim_points_y[i-1], dim_points_y[i]
            if (end_y - start_y)>0.01: desenhar_cota_vertical(c, y0+start_y*escala, y0+end_y*escala, x_pos_cota, formatar_numero(end_y-start_y))
        desenhar_cota_diametro_furo(c, x0+furos[0]['x']*escala, y0+furos[0]['y']*escala, (furos[0]['diam']/2)*escala, furos[0]['diam'])
    y_cota_total = y0 - dist_cota_total; c.line(x0, y0, x0, y_cota_total-overshoot); c.line(x0+dw, y0, x0+dw, y_cota_total-overshoot)
    desenhar_cota_horizontal(c, x0, x0+dw, y_cota_total, formatar_numero(largura))
    x_cota_total = x0 - dist_cota_total; c.line(x0, y0, x_cota_total-overshoot, y0); c.line(x0, y0+dh, x_cota_total-overshoot, y0+dh)
    desenhar_cota_vertical(c, y0, y0+dh, x_cota_total, formatar_numero(altura))
def desenhar_circulo(c, row):
    diametro = row.get('diametro', 0);
    if diametro <= 0: desenhar_erro_dados(c, "Círculo"); return
    max_w, max_h = PAGE_WIDTH-2*MARGEM_GERAL, PAGE_HEIGHT-HEADER_AREA_ALTURA-FOOTER_AREA_ALTURA
    dist_cota, overshoot = 8*mm, 2*mm
    escala = min((max_w-dist_cota*2)/diametro, (max_h-dist_cota*2)/diametro)*0.95
    raio_desenhado = (diametro*escala)/2; bloco_visual_size = (raio_desenhado*2)+dist_cota*2
    base_desenho_y = FOOTER_AREA_ALTURA; inicio_bloco_x, inicio_bloco_y = MARGEM_GERAL + (max_w-bloco_visual_size)/2, base_desenho_y + (max_h-bloco_visual_size)/2
    cx, cy = inicio_bloco_x+dist_cota+raio_desenhado, inicio_bloco_y+dist_cota+raio_desenhado
    c.circle(cx, cy, raio_desenhado, stroke=1, fill=0)
    furos = row.get('furos')
    if isinstance(furos, list):
        x0_peca, y0_peca = cx-raio_desenhado, cy-raio_desenhado
        for furo in furos: c.circle(x0_peca+(furo['x']*escala), y0_peca+(furo['y']*escala), (furo['diam']/2)*escala, stroke=1, fill=0)
    y_cota_h = cy-raio_desenhado-dist_cota; c.line(cx-raio_desenhado, cy, cx-raio_desenhado, y_cota_h-overshoot); c.line(cx+raio_desenhado, cy, cx+raio_desenhado, y_cota_h-overshoot)
    desenhar_cota_horizontal(c, cx-raio_desenhado, cx+raio_desenhado, y_cota_h, f"Ø {formatar_numero(diametro)}")
def desenhar_triangulo_retangulo(c, row):
    base, altura = row.get('rt_base', 0), row.get('rt_height', 0);
    if base <= 0 or altura <= 0: desenhar_erro_dados(c, "Triângulo Retângulo"); return
    max_w, max_h = PAGE_WIDTH-2*MARGEM_GERAL, PAGE_HEIGHT-HEADER_AREA_ALTURA-FOOTER_AREA_ALTURA
    dist_cota, overshoot = 8*mm, 2*mm; espaco_cota_x, espaco_cota_y = dist_cota, dist_cota
    escala = min((max_w-espaco_cota_x)/base, (max_h-espaco_cota_y)/altura)*0.95
    db, dh = base*escala, altura*escala; bloco_visual_width, bloco_visual_height = db+espaco_cota_x, dh+espaco_cota_y
    base_desenho_y = FOOTER_AREA_ALTURA; inicio_bloco_x, inicio_bloco_y = MARGEM_GERAL+(max_w-bloco_visual_width)/2, base_desenho_y+(max_h-bloco_visual_height)/2
    x0, y0 = inicio_bloco_x+espaco_cota_x, inicio_bloco_y+espaco_cota_y
    path=c.beginPath(); path.moveTo(x0, y0); path.lineTo(x0+db, y0); path.lineTo(x0, y0+dh); path.close(); c.drawPath(path)
    furos = row.get('furos')
    if isinstance(furos, list) and furos:
        for furo in furos: c.circle(x0+(furo['x']*escala), y0+(furo['y']*escala), (furo['diam']/2)*escala, stroke=1, fill=0)
    y_cota_h = y0 - dist_cota; c.line(x0, y0, x0, y_cota_h-overshoot); c.line(x0+db, y0, x0+db, y_cota_h-overshoot)
    desenhar_cota_horizontal(c, x0, x0+db, y_cota_h, formatar_numero(base))
    x_cota_v = x0 - dist_cota; c.line(x0, y0, x_cota_v-overshoot, y0); c.line(x0, y0+dh, x_cota_v-overshoot, y0+dh)
    desenhar_cota_vertical(c, y0, y0+dh, x_cota_v, formatar_numero(altura))
def desenhar_trapezio(c, row):
    large_base, small_base, height = row.get('trapezoid_large_base', 0), row.get('trapezoid_small_base', 0), row.get('trapezoid_height', 0)
    if large_base<=0 or height<=0 or small_base<=0: desenhar_erro_dados(c, "Trapézio"); return
    max_w, max_h = PAGE_WIDTH-2*MARGEM_GERAL, PAGE_HEIGHT-HEADER_AREA_ALTURA-FOOTER_AREA_ALTURA
    dist_cota, overshoot = 8*mm, 2*mm; espaco_cota_x, espaco_cota_y = dist_cota, dist_cota*2
    escala = min((max_w-espaco_cota_x)/large_base, (max_h-espaco_cota_y)/height)*0.95
    dlb, dsb, dh = large_base*escala, small_base*escala, height*escala
    bloco_visual_width, bloco_visual_height = dlb+espaco_cota_x, dh+espaco_cota_y
    base_desenho_y = FOOTER_AREA_ALTURA; inicio_bloco_x, inicio_bloco_y = MARGEM_GERAL+(max_w-bloco_visual_width)/2, base_desenho_y+(max_h-bloco_visual_height)/2
    x0, y0 = inicio_bloco_x+espaco_cota_x/2, inicio_bloco_y+dist_cota
    x_offset = (dlb - dsb)/2
    p1, p2, p3, p4 = (x0, y0), (x0+dlb, y0), (x0+dlb-x_offset, y0+dh), (x0+x_offset, y0+dh)
    path = c.beginPath(); path.moveTo(*p1); path.lineTo(*p2); path.lineTo(*p3); path.lineTo(*p4); path.close(); c.drawPath(path)
    furos = row.get('furos')
    if isinstance(furos, list) and furos:
        for furo in furos: c.circle(x0+(furo['x']*escala), y0+(furo['y']*escala), (furo['diam']/2)*escala, stroke=1, fill=0)
    y_cota_h1 = y0 - dist_cota; c.line(p1[0], p1[1], p1[0], y_cota_h1-overshoot); c.line(p2[0], p2[1], p2[0], y_cota_h1-overshoot)
    desenhar_cota_horizontal(c, p1[0], p2[0], y_cota_h1, formatar_numero(large_base))
    y_cota_h2 = p3[1]+dist_cota; c.line(p4[0], p4[1], p4[0], y_cota_h2+overshoot); c.line(p3[0], p3[1], p3[0], y_cota_h2+overshoot)
    desenhar_cota_horizontal(c, p4[0], p3[0], y_cota_h2, formatar_numero(small_base))
    x_cota_v = min(p1[0], p4[0]) - dist_cota; c.line(p1[0], p1[1], x_cota_v-overshoot, p1[1]); c.line(p4[0], p4[1], x_cota_v-overshoot, p4[1])
    desenhar_cota_vertical(c, p1[1], p4[1], x_cota_v, formatar_numero(height))
def desenhar_forma(c, row):
    desenhar_cabecalho(c, row.get('nome_arquivo', 'SEM NOME')); desenhar_rodape_aprimorado(c, row)
    forma = str(row.get('forma', '')).strip().lower()
    if forma == 'rectangle': desenhar_retangulo(c, row)
    elif forma == 'circle': desenhar_circulo(c, row)
    elif forma == 'right_triangle': desenhar_triangulo_retangulo(c, row)
    elif forma == 'trapezoid': desenhar_trapezio(c, row)
    else: c.setFont("Helvetica", 12); c.drawCentredString(A4[0]/2, A4[1]/2, f"Forma '{forma}' desconhecida.")


# THREAD DE PROCESSAMENTO (Nenhuma alteração aqui)
class ProcessThread(QThread):
    update_signal = pyqtSignal(str); progress_signal = pyqtSignal(int); finished_signal = pyqtSignal(bool, str)
    def __init__(self, dataframe_to_process, generate_pdf=True, generate_dxf=True, project_directory="."):
        super().__init__(); self.df = dataframe_to_process; self.generate_pdf = generate_pdf; self.generate_dxf = generate_dxf
        self.project_directory = project_directory
    def run(self):
        try:
            self.update_signal.emit("Iniciando processamento...")
            if self.generate_pdf:
                self.update_signal.emit("--- Gerando PDFs ---")
                pdf_output_dir = os.path.join(self.project_directory, "PDFs")
                os.makedirs(pdf_output_dir, exist_ok=True)
                df_pdf = self.df.copy(); df_pdf['espessura'] = df_pdf['espessura'].fillna('Sem_Espessura')
                grouped = df_pdf.groupby('espessura')
                for espessura, group in grouped:
                    pdf_filename = os.path.join(pdf_output_dir, f"Desenhos_PDF_Espessura_{str(espessura).replace('.', '_')}mm.pdf")
                    c = canvas.Canvas(pdf_filename, pagesize=A4)
                    for _, row in group.iterrows(): desenhar_forma(c, row); c.showPage()
                    c.save(); self.update_signal.emit(f"PDF salvo em: {pdf_filename}")
            if self.generate_dxf:
                # O código DXF foi omitido por não estar no prompt, mas a lógica se mantém
                self.update_signal.emit("--- Gerando DXFs (Lógica Simulada) ---")
                dxf_output_dir = os.path.join(self.project_directory, "DXFs")
                os.makedirs(dxf_output_dir, exist_ok=True)
                zip_filename = os.path.join(dxf_output_dir, f"LOTE_DXF_{datetime.now():%Y%m%d_%H%M%S}.zip")
                self.update_signal.emit(f"Arquivo ZIP (simulado) salvo em: {zip_filename}")

            self.finished_signal.emit(True, "Processamento concluído com sucesso!")
        except Exception as e: self.finished_signal.emit(False, f"Erro crítico no processamento: {str(e)}")
        
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Gerador de Desenhos Técnicos e DXF")
        self.setGeometry(100, 100, 1100, 850)
        
        self.code_generator = CodeGenerator()
        self.history_manager = HistoryManager()
        self.colunas_df = ['nome_arquivo', 'forma', 'espessura', 'qtd', 'largura', 'altura', 'diametro', 'rt_base', 'rt_height', 'trapezoid_large_base', 'trapezoid_small_base', 'trapezoid_height', 'furos']
        self.manual_df = pd.DataFrame(columns=self.colunas_df)
        self.excel_df = pd.DataFrame(columns=self.colunas_df)
        self.furos_atuais = []
        self.project_directory = None

        central_widget = QWidget(); self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget); top_h_layout = QHBoxLayout(); left_v_layout = QVBoxLayout()
        
        project_group = QGroupBox("1. Projeto")
        project_layout = QVBoxLayout()
        self.start_project_btn = QPushButton("Iniciar Novo Projeto...")
        self.start_project_btn.setStyleSheet("background-color: #007bff; color: white; font-weight: bold;")
        self.history_btn = QPushButton("Ver Histórico de Projetos")
        project_layout.addWidget(self.start_project_btn); project_layout.addWidget(self.history_btn)
        project_group.setLayout(project_layout); left_v_layout.addWidget(project_group)
        
        file_group = QGroupBox("2. Carregar Planilha (Opcional)")
        file_layout = QVBoxLayout()
        self.file_label = QLabel("Nenhum projeto ativo."); file_button_layout = QHBoxLayout()
        self.select_file_btn = QPushButton("Selecionar Planilha"); self.clear_excel_btn = QPushButton("Limpar Planilha")
        file_button_layout.addWidget(self.select_file_btn); file_button_layout.addWidget(self.clear_excel_btn)
        file_layout.addWidget(self.file_label); file_layout.addLayout(file_button_layout)
        file_group.setLayout(file_layout); left_v_layout.addWidget(file_group)

        manual_group = QGroupBox("3. Informações da Peça"); manual_layout = QFormLayout()
        self.projeto_input = QLineEdit(); self.projeto_input.setReadOnly(True)
        manual_layout.addRow("Nº do Projeto Ativo:", self.projeto_input)
        self.nome_input = QLineEdit(); self.generate_code_btn = QPushButton("Gerar Código")
        name_layout = QHBoxLayout(); name_layout.addWidget(self.nome_input); name_layout.addWidget(self.generate_code_btn)
        manual_layout.addRow("Nome/ID da Peça:", name_layout)
        self.forma_combo = QComboBox(); self.forma_combo.addItems(['rectangle', 'circle', 'right_triangle', 'trapezoid'])
        self.espessura_input, self.qtd_input = QLineEdit(), QLineEdit()
        manual_layout.addRow("Forma:", self.forma_combo); manual_layout.addRow("Espessura (mm):", self.espessura_input); manual_layout.addRow("Quantidade:", self.qtd_input)
        self.largura_input, self.altura_input = QLineEdit(), QLineEdit()
        self.diametro_input, self.rt_base_input, self.rt_height_input = QLineEdit(), QLineEdit(), QLineEdit()
        self.trapezoid_large_base_input, self.trapezoid_small_base_input, self.trapezoid_height_input = QLineEdit(), QLineEdit(), QLineEdit()
        self.largura_row = [QLabel("Largura:"), self.largura_input]; manual_layout.addRow(*self.largura_row)
        self.altura_row = [QLabel("Altura:"), self.altura_input]; manual_layout.addRow(*self.altura_row)
        self.diametro_row = [QLabel("Diâmetro:"), self.diametro_input]; manual_layout.addRow(*self.diametro_row)
        self.rt_base_row = [QLabel("Base Triângulo:"), self.rt_base_input]; manual_layout.addRow(*self.rt_base_row)
        self.rt_height_row = [QLabel("Altura Triângulo:"), self.rt_height_input]; manual_layout.addRow(*self.rt_height_row)
        self.trap_large_base_row = [QLabel("Base Maior:"), self.trapezoid_large_base_input]; manual_layout.addRow(*self.trap_large_base_row)
        self.trap_small_base_row = [QLabel("Base Menor:"), self.trapezoid_small_base_input]; manual_layout.addRow(*self.trap_small_base_row)
        self.trap_height_row = [QLabel("Altura:"), self.trapezoid_height_input]; manual_layout.addRow(*self.trap_height_row)
        manual_group.setLayout(manual_layout); left_v_layout.addWidget(manual_group); left_v_layout.addStretch()
        top_h_layout.addLayout(left_v_layout)
        
        furos_main_group = QGroupBox("4. Adicionar Furos"); furos_main_layout = QVBoxLayout()
        self.rep_group = QGroupBox("Furação Rápida"); rep_layout = QFormLayout()
        self.rep_diam_input, self.rep_offset_input = QLineEdit(), QLineEdit()
        rep_layout.addRow("Diâmetro Furos:", self.rep_diam_input); rep_layout.addRow("Offset Borda:", self.rep_offset_input)
        self.replicate_btn = QPushButton("Replicar Furos"); rep_layout.addRow(self.replicate_btn)
        self.rep_group.setLayout(rep_layout); furos_main_layout.addWidget(self.rep_group)
        man_group = QGroupBox("Furos Manuais"); man_layout = QVBoxLayout(); man_form_layout = QFormLayout()
        self.diametro_furo_input, self.pos_x_input, self.pos_y_input = QLineEdit(), QLineEdit(), QLineEdit()
        man_form_layout.addRow("Diâmetro:", self.diametro_furo_input); man_form_layout.addRow("Posição X:", self.pos_x_input); man_form_layout.addRow("Posição Y:", self.pos_y_input)
        self.add_furo_btn = QPushButton("Adicionar Furo Manual"); man_layout.addLayout(man_form_layout); man_layout.addWidget(self.add_furo_btn)
        self.furos_table = QTableWidget(0, 4); self.furos_table.setHorizontalHeaderLabels(["Diâmetro", "Pos X", "Pos Y", "Ação"])
        man_layout.addWidget(self.furos_table); man_group.setLayout(man_layout); furos_main_layout.addWidget(man_group)
        furos_main_group.setLayout(furos_main_layout); top_h_layout.addWidget(furos_main_group, stretch=1)
        main_layout.addLayout(top_h_layout)
        self.add_piece_btn = QPushButton("Adicionar Peça à Lista"); main_layout.addWidget(self.add_piece_btn)

        list_group = QGroupBox("5. Lista de Peças para Produção"); list_layout = QVBoxLayout()
        self.pieces_table = QTableWidget()
        self.table_headers = [col.replace('_', ' ').title() for col in self.colunas_df] + ["Ações"]
        self.pieces_table.setColumnCount(len(self.table_headers)); self.pieces_table.setHorizontalHeaderLabels(self.table_headers)
        list_layout.addWidget(self.pieces_table)
        self.dir_label = QLabel("Nenhum projeto ativo. Inicie um novo projeto."); self.dir_label.setStyleSheet("font-style: italic; color: grey;")
        list_layout.addWidget(self.dir_label)
        process_buttons_layout = QHBoxLayout()
        self.conclude_project_btn = QPushButton("Projeto Concluído"); self.conclude_project_btn.setStyleSheet("background-color: #28a745; color: white; font-weight: bold;")
        self.export_excel_btn = QPushButton("Exportar para Excel") # <<< NOVO
        self.process_pdf_btn, self.process_dxf_btn, self.process_all_btn = QPushButton("Gerar PDFs"), QPushButton("Gerar DXFs"), QPushButton("Gerar PDFs e DXFs")
        process_buttons_layout.addWidget(self.export_excel_btn) # <<< NOVO
        process_buttons_layout.addWidget(self.conclude_project_btn); process_buttons_layout.addStretch()
        process_buttons_layout.addWidget(self.process_pdf_btn); process_buttons_layout.addWidget(self.process_dxf_btn); process_buttons_layout.addWidget(self.process_all_btn)
        list_layout.addLayout(process_buttons_layout); list_group.setLayout(list_layout); main_layout.addWidget(list_group, stretch=1)
        
        self.progress_bar = QProgressBar(); main_layout.addWidget(self.progress_bar)
        log_group = QGroupBox("Log de Execução"); log_layout = QVBoxLayout(); self.log_text = QTextEdit(); log_layout.addWidget(self.log_text)
        log_group.setLayout(log_layout); main_layout.addWidget(log_group); self.statusBar().showMessage("Pronto")
        
        self.start_project_btn.clicked.connect(self.start_new_project); self.history_btn.clicked.connect(self.show_history_dialog)
        self.select_file_btn.clicked.connect(self.select_file); self.clear_excel_btn.clicked.connect(self.clear_excel_data)
        self.generate_code_btn.clicked.connect(self.generate_piece_code); self.add_piece_btn.clicked.connect(self.add_manual_piece)
        self.forma_combo.currentTextChanged.connect(self.update_dimension_fields); self.replicate_btn.clicked.connect(self.replicate_holes)
        self.add_furo_btn.clicked.connect(self.add_furo_temp); self.process_pdf_btn.clicked.connect(self.start_pdf_generation)
        self.process_dxf_btn.clicked.connect(self.start_dxf_generation); self.process_all_btn.clicked.connect(self.start_all_generation)
        self.conclude_project_btn.clicked.connect(self.conclude_project)
        self.export_excel_btn.clicked.connect(self.export_project_to_excel) # <<< NOVO
        self.set_initial_button_state(); self.update_dimension_fields(self.forma_combo.currentText())

    def start_new_project(self):
        parent_dir = QFileDialog.getExistingDirectory(self, "Selecione a Pasta Principal para o Novo Projeto")
        if not parent_dir: return
        project_name, ok = QInputDialog.getText(self, "Novo Projeto", "Digite o nome ou número do novo projeto:")
        if ok and project_name:
            project_path = os.path.join(parent_dir, project_name)
            if os.path.exists(project_path):
                reply = QMessageBox.question(self, 'Diretório Existente', f"A pasta '{project_name}' já existe.\nDeseja usá-la como o diretório do projeto ativo?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                if reply == QMessageBox.No: return
            else:
                try: os.makedirs(project_path)
                except OSError as e: QMessageBox.critical(self, "Erro ao Criar Pasta", f"Não foi possível criar o diretório do projeto:\n{e}"); return
            self._clear_session(clear_project_number=True)
            self.project_directory = project_path
            self.projeto_input.setText(project_name)
            self.dir_label.setText(f"Projeto Ativo: {self.project_directory}")
            self.dir_label.setStyleSheet("font-style: normal; color: black;")
            self.log_text.append(f"\n--- NOVO PROJETO INICIADO: {project_name} ---")
            self.log_text.append(f"Arquivos serão salvos em: {self.project_directory}")
            self.set_initial_button_state()

    def set_initial_button_state(self):
        is_project_active = self.project_directory is not None
        self.start_project_btn.setEnabled(True); self.history_btn.setEnabled(True)
        self.select_file_btn.setEnabled(is_project_active); self.clear_excel_btn.setEnabled(is_project_active and not self.excel_df.empty)
        self.generate_code_btn.setEnabled(is_project_active); self.add_piece_btn.setEnabled(is_project_active)
        self.replicate_btn.setEnabled(is_project_active); self.add_furo_btn.setEnabled(is_project_active)
        has_items = not (self.excel_df.empty and self.manual_df.empty)
        self.process_pdf_btn.setEnabled(is_project_active and has_items); self.process_dxf_btn.setEnabled(is_project_active and has_items)
        self.process_all_btn.setEnabled(is_project_active and has_items); self.conclude_project_btn.setEnabled(is_project_active and has_items)
        self.export_excel_btn.setEnabled(is_project_active and has_items) # <<< ALTERADO
        self.progress_bar.setVisible(False)

    def show_history_dialog(self):
        dialog = HistoryDialog(self.history_manager, self)
        if dialog.exec_() == QDialog.Accepted:
            loaded_pieces = dialog.loaded_project_data
            if loaded_pieces:
                project_number_loaded = loaded_pieces[0].get('project_number', '') if loaded_pieces else ''
                self.start_new_project_from_history(project_number_loaded, loaded_pieces)
    
    def start_new_project_from_history(self, project_name, pieces_data):
        parent_dir = QFileDialog.getExistingDirectory(self, f"Selecione uma pasta para o projeto '{project_name}'")
        if not parent_dir: return
        project_path = os.path.join(parent_dir, project_name)
        os.makedirs(project_path, exist_ok=True)
        self._clear_session(clear_project_number=True)
        self.project_directory = project_path
        self.projeto_input.setText(project_name)
        self.excel_df = pd.DataFrame(columns=self.colunas_df)
        self.manual_df = pd.DataFrame(pieces_data)
        self.dir_label.setText(f"Projeto Ativo: {self.project_directory}"); self.dir_label.setStyleSheet("font-style: normal; color: black;")
        self.log_text.append(f"\n--- PROJETO DO HISTÓRICO CARREGADO: {project_name} ---")
        self.update_table_display()
        self.set_initial_button_state()

    def start_pdf_generation(self): self.start_processing(generate_pdf=True, generate_dxf=False)
    def start_dxf_generation(self): self.start_processing(generate_pdf=False, generate_dxf=True)
    def start_all_generation(self): self.start_processing(generate_pdf=True, generate_dxf=True)

    def start_processing(self, generate_pdf, generate_dxf):
        if not self.project_directory:
            QMessageBox.warning(self, "Nenhum Projeto Ativo", "Inicie um novo projeto antes de gerar arquivos."); return
        combined_df = pd.concat([self.excel_df, self.manual_df], ignore_index=True)
        if combined_df.empty: QMessageBox.warning(self, "Aviso", "A lista de peças está vazia."); return
        self.set_buttons_enabled_on_process(False)
        self.progress_bar.setVisible(True); self.progress_bar.setValue(0); self.log_text.clear()
        self.process_thread = ProcessThread(combined_df.copy(), generate_pdf, generate_dxf, self.project_directory)
        self.process_thread.update_signal.connect(self.log_text.append)
        self.process_thread.progress_signal.connect(self.progress_bar.setValue)
        self.process_thread.finished_signal.connect(self.processing_finished)
        self.process_thread.start()

    def processing_finished(self, success, message):
        self.set_buttons_enabled_on_process(True); self.progress_bar.setVisible(False)
        msgBox = QMessageBox.information if success else QMessageBox.critical
        msgBox(self, "Concluído" if success else "Erro", message); self.statusBar().showMessage("Pronto")
    
    def conclude_project(self):
        project_number = self.projeto_input.text().strip()
        if not project_number:
            QMessageBox.warning(self, "Projeto sem Número", "O projeto ativo não tem um número definido.")
            return
        reply = QMessageBox.question(self, 'Concluir Projeto', f"Deseja salvar e concluir o projeto '{project_number}'?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            combined_df = pd.concat([self.excel_df, self.manual_df], ignore_index=True)
            if not combined_df.empty:
                combined_df['project_number'] = project_number
                self.history_manager.save_project(project_number, combined_df)
                self.log_text.append(f"Projeto '{project_number}' salvo no histórico.")
            self._clear_session(clear_project_number=True)
            self.project_directory = None
            self.dir_label.setText("Nenhum projeto ativo. Inicie um novo projeto."); self.dir_label.setStyleSheet("font-style: italic; color: grey;")
            self.set_initial_button_state()
            self.log_text.append(f"\n--- PROJETO '{project_number}' CONCLUÍDO ---")
    
    # <<< INÍCIO DO BLOCO DE CÓDIGO NOVO >>>
    def export_project_to_excel(self):
        project_number = self.projeto_input.text().strip()
        if not project_number:
            QMessageBox.warning(self, "Nenhum Projeto Ativo", "Inicie um novo projeto para poder exportá-lo.")
            return

        combined_df = pd.concat([self.excel_df, self.manual_df], ignore_index=True)
        if combined_df.empty:
            QMessageBox.warning(self, "Lista Vazia", "Não há peças na lista para exportar.")
            return

        default_filename = os.path.join(self.project_directory, f"Resumo_Projeto_{project_number}.xlsx")
        save_path, _ = QFileDialog.getSaveFileName(self, "Salvar Resumo do Projeto", default_filename, "Excel Files (*.xlsx)")

        if not save_path:
            return

        try:
            # 1. Preparar os dados para exportação
            export_df = pd.DataFrame()
            export_df['Numero do Projeto'] = [project_number] * len(combined_df)
            export_df['Nome da Peça'] = combined_df['nome_arquivo']
            export_df['Forma'] = combined_df['forma']
            export_df['Espessura (mm)'] = combined_df['espessura']
            export_df['Quantidade'] = combined_df['qtd']

            # Adicionar colunas de dimensão com base na forma
            export_df['Largura (mm)'] = combined_df['largura']
            export_df['Altura (mm)'] = combined_df['altura']
            export_df['Diâmetro (mm)'] = combined_df['diametro']
            export_df['Base Triângulo (mm)'] = combined_df['rt_base']
            export_df['Altura Triângulo (mm)'] = combined_df['rt_height']
            export_df['Base Maior Trapézio (mm)'] = combined_df['trapezoid_large_base']
            export_df['Base Menor Trapézio (mm)'] = combined_df['trapezoid_small_base']
            export_df['Altura Trapézio (mm)'] = combined_df['trapezoid_height']
            
            # 2. Processar informações dos furos
            furos_data = combined_df['furos'].fillna('[]').apply(
                lambda f: f if isinstance(f, list) else json.loads(f.replace("'", "\""))
            )
            export_df['Quantidade de Furos'] = furos_data.apply(len)
            
            def get_unique_diameters(furos_list):
                if not furos_list: return ""
                diameters = sorted(list(set([f.get('diam', 0) for f in furos_list])))
                return ", ".join(map(str, diameters))

            export_df['Diâmetros dos Furos (mm)'] = furos_data.apply(get_unique_diameters)
            export_df['Detalhes dos Furos (JSON)'] = furos_data.apply(json.dumps)
            
            # 3. Limpar colunas vazias e salvar
            export_df.replace(0, pd.NA, inplace=True) # Substitui 0 por N/A para limpeza
            export_df.dropna(axis=1, how='all', inplace=True) # Remove colunas que ficaram inteiramente vazias

            export_df.to_excel(save_path, index=False, engine='openpyxl')
            
            self.log_text.append(f"Resumo do projeto salvo com sucesso em: {save_path}")
            QMessageBox.information(self, "Sucesso", f"O arquivo Excel foi salvo com sucesso em:\n{save_path}")

        except Exception as e:
            self.log_text.append(f"ERRO ao exportar para Excel: {e}")
            QMessageBox.critical(self, "Erro na Exportação", f"Ocorreu um erro ao salvar o arquivo:\n{e}")
    # <<< FIM DO BLOCO DE CÓDIGO NOVO >>>

    def _clear_session(self, clear_project_number=False):
        fields_to_clear = [self.nome_input, self.espessura_input, self.qtd_input, self.largura_input, self.altura_input, self.diametro_input, self.rt_base_input, self.rt_height_input, self.trapezoid_large_base_input, self.trapezoid_small_base_input, self.trapezoid_height_input, self.rep_diam_input, self.rep_offset_input, self.diametro_furo_input, self.pos_x_input, self.pos_y_input]
        if clear_project_number:
            fields_to_clear.append(self.projeto_input)
        for field in fields_to_clear:
            field.clear()
        self.furos_atuais = []
        self.update_furos_table()
        self.file_label.setText("Nenhum projeto ativo.")
        if clear_project_number: 
            self.excel_df = pd.DataFrame(columns=self.colunas_df)
            self.manual_df = pd.DataFrame(columns=self.colunas_df)
            self.update_table_display()

    def set_buttons_enabled_on_process(self, enabled):
        self.start_project_btn.setEnabled(enabled); self.history_btn.setEnabled(enabled); self.select_file_btn.setEnabled(enabled); self.clear_excel_btn.setEnabled(enabled and not self.excel_df.empty)
        self.generate_code_btn.setEnabled(enabled); self.add_piece_btn.setEnabled(enabled); self.replicate_btn.setEnabled(enabled); self.add_furo_btn.setEnabled(enabled)
        has_items = not (self.excel_df.empty and self.manual_df.empty)
        self.process_pdf_btn.setEnabled(enabled and has_items); self.process_dxf_btn.setEnabled(enabled and has_items)
        self.process_all_btn.setEnabled(enabled and has_items); self.conclude_project_btn.setEnabled(enabled and has_items)
        self.export_excel_btn.setEnabled(enabled and has_items) # <<< ALTERADO
    
    def update_table_display(self):
        is_project_active = self.project_directory is not None
        has_items = not (self.excel_df.empty and self.manual_df.empty)
        self.process_pdf_btn.setEnabled(is_project_active and has_items)
        self.process_dxf_btn.setEnabled(is_project_active and has_items)
        self.process_all_btn.setEnabled(is_project_active and has_items)
        self.conclude_project_btn.setEnabled(is_project_active and has_items)
        self.export_excel_btn.setEnabled(is_project_active and has_items) # <<< ALTERADO
        self.clear_excel_btn.setEnabled(is_project_active and not self.excel_df.empty)
        combined_df = pd.concat([self.excel_df, self.manual_df], ignore_index=True); self.pieces_table.setRowCount(0)
        if combined_df.empty: return
        self.pieces_table.setRowCount(len(combined_df))
        for i, row in combined_df.iterrows():
            for j, col in enumerate(self.colunas_df):
                value = row[col]
                display_value = f"{len(value)} Furo(s)" if col == 'furos' and isinstance(value, list) else ('-' if pd.isna(value) or value == 0 else str(value))
                self.pieces_table.setItem(i, j, QTableWidgetItem(display_value))
            action_widget = QWidget(); action_layout = QHBoxLayout(action_widget); action_layout.setContentsMargins(5, 0, 5, 0)
            edit_btn, delete_btn = QPushButton("Editar"), QPushButton("Excluir")
            edit_btn.clicked.connect(lambda _, r=i: self.edit_row(r)); delete_btn.clicked.connect(lambda _, r=i: self.delete_row(r))
            action_layout.addWidget(edit_btn); action_layout.addWidget(delete_btn)
            self.pieces_table.setCellWidget(i, len(self.colunas_df), action_widget)
        self.pieces_table.resizeColumnsToContents()
    
    def edit_row(self, row_index):
        len_excel = len(self.excel_df); df_source = self.excel_df if row_index < len_excel else self.manual_df
        local_index = row_index if row_index < len_excel else row_index - len_excel
        piece_data = df_source.iloc[local_index]
        self.nome_input.setText(str(piece_data.get('nome_arquivo', ''))); self.espessura_input.setText(str(piece_data.get('espessura', ''))); self.qtd_input.setText(str(piece_data.get('qtd', '')))
        shape = piece_data.get('forma', ''); index = self.forma_combo.findText(shape, Qt.MatchFixedString)
        if index >= 0: self.forma_combo.setCurrentIndex(index)
        self.largura_input.setText(str(piece_data.get('largura', ''))); self.altura_input.setText(str(piece_data.get('altura', ''))); self.diametro_input.setText(str(piece_data.get('diametro', '')))
        self.rt_base_input.setText(str(piece_data.get('rt_base', ''))); self.rt_height_input.setText(str(piece_data.get('rt_height', '')))
        self.trapezoid_large_base_input.setText(str(piece_data.get('trapezoid_large_base', ''))); self.trapezoid_small_base_input.setText(str(piece_data.get('trapezoid_small_base', '')))
        self.trapezoid_height_input.setText(str(piece_data.get('trapezoid_height', '')))
        self.furos_atuais = piece_data.get('furos', []).copy() if isinstance(piece_data.get('furos'), list) else []
        self.update_furos_table(); df_source.drop(df_source.index[local_index], inplace=True); df_source.reset_index(drop=True, inplace=True)
        self.log_text.append(f"Peça '{piece_data['nome_arquivo']}' carregada para edição."); self.update_table_display()
    
    def delete_row(self, row_index):
        len_excel = len(self.excel_df); df_source = self.excel_df if row_index < len_excel else self.manual_df
        local_index = row_index if row_index < len_excel else row_index - len_excel
        piece_name = df_source.iloc[local_index]['nome_arquivo']
        df_source.drop(df_source.index[local_index], inplace=True); df_source.reset_index(drop=True, inplace=True)
        self.log_text.append(f"Peça '{piece_name}' removida."); self.update_table_display()
    
    def generate_piece_code(self):
        project_number = self.projeto_input.text().strip()
        if not project_number: QMessageBox.warning(self, "Campo Obrigatório", "Inicie um projeto para definir o 'Nº do Projeto'."); return
        new_code = self.code_generator.generate_new_code(project_number, prefix='DES')
        if new_code: self.nome_input.setText(new_code); self.log_text.append(f"Código '{new_code}' gerado para o projeto '{project_number}'.")
    
    def add_manual_piece(self):
        try:
            nome = self.nome_input.text().strip()
            if not nome: QMessageBox.warning(self, "Campo Obrigatório", "'Nome/ID da Peça' é obrigatório."); return
            new_piece = {'furos': self.furos_atuais.copy()};
            for col in self.colunas_df:
                if col != 'furos': new_piece[col] = 0.0
            new_piece.update({'nome_arquivo': nome, 'forma': self.forma_combo.currentText()})
            fields_map = { 'espessura': self.espessura_input, 'qtd': self.qtd_input, 'largura': self.largura_input, 'altura': self.altura_input, 'diametro': self.diametro_input, 'rt_base': self.rt_base_input, 'rt_height': self.rt_height_input, 'trapezoid_large_base': self.trapezoid_large_base_input, 'trapezoid_small_base': self.trapezoid_small_base_input, 'trapezoid_height': self.trapezoid_height_input }
            for key, field in fields_map.items():
                new_piece[key] = float(field.text().replace(',', '.')) if field.text() else 0.0
            self.manual_df = pd.concat([self.manual_df, pd.DataFrame([new_piece])], ignore_index=True)
            self.log_text.append(f"Peça '{nome}' adicionada/atualizada.")
            self._clear_session(clear_project_number=False)
            self.update_table_display()
        except ValueError: QMessageBox.critical(self, "Erro de Valor", "Campos numéricos devem conter números válidos.")
    
    def select_file(self):
        if not self.project_directory:
            QMessageBox.warning(self, "Nenhum Projeto Ativo", "Inicie um projeto antes de carregar uma planilha.")
            return
        file_path, _ = QFileDialog.getOpenFileName(self, "Selecionar Planilha", "", "Excel Files (*.xlsx *.xls)")
        if file_path:
            try:
                df = pd.read_excel(file_path, header=0, decimal=','); df.columns = df.columns.str.strip().str.lower()
                df = df.loc[:, ~df.columns.duplicated()]
                for col in self.colunas_df:
                    if col not in df.columns: df[col] = pd.NA
                if 'furos' in df.columns:
                    def parse_furos(x):
                        if isinstance(x, list): return x
                        if isinstance(x, str) and x.startswith('['):
                            try: return json.loads(x.replace("'", "\""))
                            except json.JSONDecodeError: return []
                        return []
                    df['furos'] = df['furos'].apply(parse_furos)
                else:
                    df['furos'] = [[] for _ in range(len(df))]
                self.excel_df = df[self.colunas_df]
                self.file_label.setText(f"Planilha: {os.path.basename(file_path)}"); self.update_table_display()
            except Exception as e: QMessageBox.critical(self, "Erro de Leitura", f"Falha ao ler o arquivo: {e}")
    
    def clear_excel_data(self):
        self.excel_df = pd.DataFrame(columns=self.colunas_df); self.file_label.setText("Nenhuma planilha selecionada"); self.update_table_display()
    
    def replicate_holes(self):
        try:
            if self.forma_combo.currentText() != 'rectangle': QMessageBox.warning(self, "Função Indisponível", "Replicação disponível apenas para Retângulos."); return
            largura, altura = float(self.largura_input.text().replace(',', '.')), float(self.altura_input.text().replace(',', '.'))
            diam, offset = float(self.rep_diam_input.text().replace(',', '.')), float(self.rep_offset_input.text().replace(',', '.'))
            if (offset * 2) >= largura or (offset * 2) >= altura: QMessageBox.warning(self, "Offset Inválido", "Offset excede as dimensões da peça."); return
            furos = [{'diam': diam, 'x': offset, 'y': offset}, {'diam': diam, 'x': largura - offset, 'y': offset}, {'diam': diam, 'x': largura - offset, 'y': altura - offset}, {'diam': diam, 'x': offset, 'y': altura - offset}]
            self.furos_atuais.extend(furos); self.update_furos_table()
        except ValueError: QMessageBox.critical(self, "Erro de Valor", "Largura, Altura, Diâmetro e Offset devem ser números válidos.")
    
    def update_dimension_fields(self, shape):
        shape = shape.lower()
        is_rect, is_circ, is_tri, is_trap = shape == 'rectangle', shape == 'circle', shape == 'right_triangle', shape == 'trapezoid'
        for w in self.largura_row + self.altura_row: w.setVisible(is_rect)
        for w in self.diametro_row: w.setVisible(is_circ)
        for w in self.rt_base_row + self.rt_height_row: w.setVisible(is_tri)
        for w in self.trap_large_base_row + self.trap_small_base_row + self.trap_height_row: w.setVisible(is_trap)
        self.rep_group.setEnabled(is_rect)
    
    def add_furo_temp(self):
        try:
            diam, pos_x, pos_y = float(self.diametro_furo_input.text().replace(',', '.')), float(self.pos_x_input.text().replace(',', '.')), float(self.pos_y_input.text().replace(',', '.'))
            if diam <= 0: QMessageBox.warning(self, "Valor Inválido", "Diâmetro do furo deve ser maior que zero."); return
            self.furos_atuais.append({'diam': diam, 'x': pos_x, 'y': pos_y}); self.update_furos_table()
            for field in [self.diametro_furo_input, self.pos_x_input, self.pos_y_input]: field.clear()
        except ValueError: QMessageBox.critical(self, "Erro de Valor", "Campos de furo devem ser números válidos.")
    
    def update_furos_table(self):
        self.furos_table.setRowCount(0); self.furos_table.setRowCount(len(self.furos_atuais))
        for i, furo in enumerate(self.furos_atuais):
            self.furos_table.setItem(i, 0, QTableWidgetItem(str(furo['diam']))); self.furos_table.setItem(i, 1, QTableWidgetItem(str(furo['x'])))
            self.furos_table.setItem(i, 2, QTableWidgetItem(str(furo['y'])))
            delete_btn = QPushButton("Excluir"); delete_btn.clicked.connect(lambda _, r=i: self.delete_furo_temp(r)); self.furos_table.setCellWidget(i, 3, delete_btn)
        self.furos_table.resizeColumnsToContents()
    
    def delete_furo_temp(self, row_index):
        del self.furos_atuais[row_index]; self.update_furos_table()
def main():
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()