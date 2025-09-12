# gerador_app.py

import pandas as pd
import math
import sys
import os
import zipfile
from datetime import datetime
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                             QHBoxLayout, QPushButton, QLabel, QTextEdit, 
                             QFileDialog, QProgressBar, QMessageBox, QGroupBox,
                             QFormLayout, QLineEdit, QComboBox, QTableWidget, QTableWidgetItem)
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from openpyxl import load_workbook, Workbook

# Importa a nossa nova engine de DXF
import dxf_engine

# ==============================================
# CONSTANTES GLOBAIS DE LAYOUT DA PÁGINA
# ==============================================
PAGE_WIDTH, PAGE_HEIGHT = A4
MARGEM_GERAL = 15 * mm
HEADER_AREA_ALTURA = 25 * mm
FOOTER_AREA_ALTURA = 25 * mm 

# ==============================================
# CLASSE PARA GERENCIAR CÓDIGOS ÚNICOS
# ==============================================
class CodeGenerator:
    def __init__(self, db_path="codigo_database.xlsx"):
        self.db_path = db_path
        self.code_column_name = 'Codigo Unico'
        self.timestamp_column_name = 'Data de Registro'
        self.project_column_name = 'Projeto'
        
        self.existing_codes = set()
        self.last_code_number = 0
        self._load_database()

    def _load_database(self):
        try:
            df = pd.read_excel(self.db_path)
            if self.code_column_name in df.columns:
                df[self.code_column_name] = df[self.code_column_name].astype(str)
                self.existing_codes = set(df[self.code_column_name].tolist())
                
                max_num = 0
                for code in self.existing_codes:
                    if code.upper().startswith('DES'):
                        try:
                            num = int(code[3:])
                            if num > max_num: max_num = num
                        except (ValueError, TypeError): continue
                self.last_code_number = max_num
        except FileNotFoundError:
            print(f"Arquivo '{self.db_path}' não encontrado. Um novo será criado.")
        except Exception as e:
            QMessageBox.warning(None, "Erro de Leitura", f"Não foi possível ler o banco de dados de códigos: {e}")

    def _append_to_database(self, new_code, project_number):
        timestamp = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        new_row = [new_code, timestamp, project_number]

        try:
            wb = load_workbook(self.db_path)
            sheet = wb.active
        except FileNotFoundError:
            wb = Workbook()
            sheet = wb.active
            sheet.append([self.code_column_name, self.timestamp_column_name, self.project_column_name])
        
        sheet.append(new_row)
        
        try:
            wb.save(self.db_path)
        except PermissionError:
             QMessageBox.critical(None, "Erro de Permissão", f"Não foi possível salvar o novo código em '{self.db_path}'.\n\nPor favor, feche a planilha e tente novamente.")
             return False
        except Exception as e:
            QMessageBox.critical(None, "Erro ao Salvar", f"Ocorreu um erro ao salvar o banco de códigos: {e}")
            return False
        return True

    def generate_new_code(self, project_number, prefix='DES'):
        self._load_database()
        current_number = self.last_code_number + 1
        new_code = f"{prefix}{current_number}"
        
        while new_code in self.existing_codes:
            current_number += 1
            new_code = f"{prefix}{current_number}"
        
        if self._append_to_database(new_code, project_number):
            self.existing_codes.add(new_code)
            self.last_code_number = current_number
            return new_code
        else:
            return None

# ==============================================
# FUNÇÕES DE DESENHO E RELATÓRIO PDF
# ==============================================
def desenhar_cabecalho(c, nome_arquivo):
    c.setFont("Helvetica-Bold", 14)
    y_pos_texto = PAGE_HEIGHT - HEADER_AREA_ALTURA + (10 * mm)
    c.drawCentredString(PAGE_WIDTH / 2, y_pos_texto, f"Desenho da Peça: {nome_arquivo}")
    y_pos_linha = PAGE_HEIGHT - HEADER_AREA_ALTURA
    c.line(MARGEM_GERAL, y_pos_linha, PAGE_WIDTH - MARGEM_GERAL, y_pos_linha)

def desenhar_rodape_aprimorado(c, row):
    largura_total = PAGE_WIDTH - 2 * MARGEM_GERAL
    altura_bloco = 12 * mm
    y_rodape = FOOTER_AREA_ALTURA - altura_bloco - (5 * mm)
    c.setStrokeColorRGB(0, 0, 0)
    c.rect(MARGEM_GERAL, y_rodape, largura_total, altura_bloco)
    coluna1_x, coluna2_x, coluna3_x = MARGEM_GERAL, MARGEM_GERAL + largura_total * 0.60, MARGEM_GERAL + largura_total * 0.80
    c.line(coluna2_x, y_rodape, coluna2_x, y_rodape + altura_bloco); c.line(coluna3_x, y_rodape, coluna3_x, y_rodape + altura_bloco)
    y_titulo, y_valor = y_rodape + altura_bloco - 4*mm, y_rodape + 4*mm
    c.setFont("Helvetica", 7); c.drawCentredString(coluna1_x + (coluna2_x - coluna1_x)/2, y_titulo, "NOME DA PEÇA / IDENTIFICADOR")
    c.setFont("Helvetica-Bold", 10); c.drawCentredString(coluna1_x + (coluna2_x - coluna1_x)/2, y_valor, str(row.get('nome_arquivo', 'N/A')))
    c.setFont("Helvetica", 7); c.drawCentredString(coluna2_x + (coluna3_x - coluna2_x)/2, y_titulo, "ESPESSURA")
    c.setFont("Helvetica-Bold", 10); c.drawCentredString(coluna2_x + (coluna3_x - coluna2_x)/2, y_valor, f"{formatar_numero(row.get('espessura', 0))} mm")
    c.setFont("Helvetica", 7); c.drawCentredString(coluna3_x + (PAGE_WIDTH - MARGEM_GERAL - coluna3_x)/2, y_titulo, "QUANTIDADE")
    c.setFont("Helvetica-Bold", 10); c.drawCentredString(coluna3_x + (PAGE_WIDTH - MARGEM_GERAL - coluna3_x)/2, y_valor, formatar_numero(row.get('qtd', 0)))

def desenhar_erro_dados(c, forma):
    c.setFont("Helvetica-Bold", 14); c.drawCentredString(A4[0]/2, A4[1]/2, f"Dados inválidos para a forma: '{forma}'")

def formatar_numero(valor):
    if valor == int(valor): return str(int(valor))
    return str(valor)

def desenhar_cota_horizontal(c, x1, x2, y, texto):
    FONT_NAME, FONT_SIZE = "Helvetica", 10
    c.setFont(FONT_NAME, FONT_SIZE)
    c.line(x1, y, x2, y)
    tick_len = 2 * mm
    c.line(x1, y - tick_len/2, x1 + tick_len/2, y + tick_len/2)
    c.line(x2, y - tick_len/2, x2 - tick_len/2, y + tick_len/2)
    largura_texto = c.stringWidth(texto, FONT_NAME, FONT_SIZE)
    centro_x, y_texto = (x1 + x2) / 2, y + 1*mm
    padding = 1.5 * mm
    gap_inicio, gap_fim = centro_x - (largura_texto / 2) - padding, centro_x + (largura_texto / 2) + padding
    c.line(x1, y, gap_inicio, y)
    c.line(gap_fim, y, x2, y)
    c.drawCentredString(centro_x, y_texto, texto)

def desenhar_cota_vertical(c, y1, y2, x, texto):
    FONT_NAME, FONT_SIZE = "Helvetica", 10
    c.setFont(FONT_NAME, FONT_SIZE)
    c.line(x, y1, x, y2)
    tick_len = 2 * mm
    c.line(x - tick_len/2, y1, x + tick_len/2, y1 + tick_len/2)
    c.line(x - tick_len/2, y2, x + tick_len/2, y2 - tick_len/2)
    largura_texto, altura_texto = c.stringWidth(texto, FONT_NAME, FONT_SIZE), FONT_SIZE
    centro_y = (y1 + y2) / 2
    c.saveState()
    c.translate(x, centro_y)
    c.rotate(90)
    c.setFillColorRGB(1, 1, 1)
    c.rect(-largura_texto/2 - 1*mm, -altura_texto/2, largura_texto + 2*mm, altura_texto, stroke=0, fill=1)
    c.setFillColorRGB(0, 0, 0)
    c.drawCentredString(0, -1.5*mm, texto)
    c.restoreState()

def desenhar_cota_diametro_furo(c, x, y, raio, diametro_real):
    c.saveState()
    c.setFont("Helvetica", 10)
    texto = f"Ø {formatar_numero(diametro_real)}"
    largura_texto = c.stringWidth(texto, "Helvetica", 10)
    p_borda_x, p_borda_y = x + raio * 0.7071, y + raio * 0.7071
    p_meio_x, p_meio_y = p_borda_x + 4 * mm, p_borda_y + 4 * mm
    p_final_x = p_meio_x + largura_texto + 2*mm
    path = c.beginPath()
    path.moveTo(p_borda_x, p_borda_y); path.lineTo(p_meio_x, p_meio_y); path.lineTo(p_final_x, p_meio_y)
    c.drawPath(path)
    c.drawString(p_meio_x + 1*mm, p_meio_y + 1*mm, texto)
    c.restoreState()

def desenhar_retangulo(c, row):
    largura, altura = row.get('largura', 0), row.get('altura', 0)
    if largura <= 0 or altura <= 0: desenhar_erro_dados(c, "Retângulo"); return
    max_w = PAGE_WIDTH - 2 * MARGEM_GERAL
    max_h = PAGE_HEIGHT - HEADER_AREA_ALTURA - FOOTER_AREA_ALTURA
    dist_cota_furo, dist_cota_total, overshoot = 8 * mm, 16 * mm, 2 * mm
    espaco_cota_x, espaco_cota_y = dist_cota_total, dist_cota_total
    escala = min((max_w - espaco_cota_x) / largura, (max_h - espaco_cota_y) / altura) * 0.95
    dw, dh = largura * escala, altura * escala
    bloco_visual_width, bloco_visual_height = dw + espaco_cota_x, dh + espaco_cota_y
    base_desenho_y = FOOTER_AREA_ALTURA
    inicio_bloco_x, inicio_bloco_y = MARGEM_GERAL + (max_w - bloco_visual_width) / 2, base_desenho_y + (max_h - bloco_visual_height) / 2
    x0, y0 = inicio_bloco_x + espaco_cota_x, inicio_bloco_y + espaco_cota_y
    c.rect(x0, y0, dw, dh)
    furos = row.get('furos')
    if isinstance(furos, list) and furos:
        for furo in furos: c.circle(x0 + (furo['x'] * escala), y0 + (furo['y'] * escala), (furo['diam'] / 2) * escala, stroke=1, fill=0)
        y_pos_cota = y0 - dist_cota_furo
        unique_x = sorted(list(set(f['x'] for f in furos))); dim_points_x = [0] + unique_x + [largura]
        for x_real in dim_points_x: c.line(x0 + x_real * escala, y0, x0 + x_real * escala, y_pos_cota - overshoot)
        for i in range(1, len(dim_points_x)):
            start_x, end_x = dim_points_x[i-1], dim_points_x[i]
            if (end_x - start_x) > 0.01: desenhar_cota_horizontal(c, x0 + start_x * escala, x0 + end_x * escala, y_pos_cota, formatar_numero(end_x - start_x))
        x_pos_cota = x0 - dist_cota_furo
        unique_y = sorted(list(set(f['y'] for f in furos))); dim_points_y = [0] + unique_y + [altura]
        for y_real in dim_points_y: c.line(x0, y0 + y_real * escala, x_pos_cota - overshoot, y0 + y_real * escala)
        for i in range(1, len(dim_points_y)):
            start_y, end_y = dim_points_y[i-1], dim_points_y[i]
            if (end_y - start_y) > 0.01: desenhar_cota_vertical(c, y0 + start_y * escala, y0 + end_y * escala, x_pos_cota, formatar_numero(end_y - start_y))
        desenhar_cota_diametro_furo(c, x0 + furos[0]['x'] * escala, y0 + furos[0]['y'] * escala, (furos[0]['diam'] / 2) * escala, furos[0]['diam'])
    y_cota_total = y0 - dist_cota_total
    c.line(x0, y0, x0, y_cota_total - overshoot); c.line(x0 + dw, y0, x0 + dw, y_cota_total - overshoot)
    desenhar_cota_horizontal(c, x0, x0 + dw, y_cota_total, formatar_numero(largura))
    x_cota_total = x0 - dist_cota_total
    c.line(x0, y0, x_cota_total - overshoot, y0); c.line(x0, y0 + dh, x_cota_total - overshoot, y0 + dh)
    desenhar_cota_vertical(c, y0, y0 + dh, x_cota_total, formatar_numero(altura))

def desenhar_circulo(c, row):
    # Esta função pode ser aprimorada com a mesma lógica de centralização e cotas do retângulo.
    desenhar_erro_dados(c, "Círculo (Desenho não implementado)")

def desenhar_triangulo_retangulo(c, row):
    # Esta função pode ser aprimorada com a mesma lógica de centralização e cotas do retângulo.
    desenhar_erro_dados(c, "Triângulo (Desenho não implementado)")

def desenhar_forma(c, row):
    desenhar_cabecalho(c, row.get('nome_arquivo', 'SEM NOME'))
    desenhar_rodape_aprimorado(c, row)
    forma = str(row.get('forma', '')).strip().lower()
    if forma == 'rectangle':
        desenhar_retangulo(c, row)
    elif forma == 'circle':
        desenhar_circulo(c, row)
    elif forma == 'right_triangle':
        desenhar_triangulo_retangulo(c, row)
    else:
        c.setFont("Helvetica", 12); c.drawCentredString(A4[0]/2, A4[1]/2, f"Forma '{forma}' desconhecida.")

# ==============================================
# THREAD DE PROCESSAMENTO
# ==============================================
class ProcessThread(QThread):
    update_signal = pyqtSignal(str)
    progress_signal = pyqtSignal(int)
    finished_signal = pyqtSignal(bool, str)
    
    def __init__(self, dataframe_to_process, generate_pdf=True, generate_dxf=True):
        super().__init__()
        self.df = dataframe_to_process
        self.generate_pdf = generate_pdf
        self.generate_dxf = generate_dxf
        
    def run(self):
        try:
            self.update_signal.emit("Iniciando processamento...")
            
            if self.generate_pdf:
                self.update_signal.emit("--- Gerando PDFs ---")
                df_pdf = self.df.copy()
                df_pdf['espessura'] = df_pdf['espessura'].fillna('Sem_Espessura')
                grouped = df_pdf.groupby('espessura')
                for espessura, group in grouped:
                    pdf_filename = f"Desenhos_PDF_Espessura_{str(espessura).replace('.', '_')}mm.pdf"
                    c = canvas.Canvas(pdf_filename, pagesize=A4)
                    for _, row in group.iterrows(): desenhar_forma(c, row); c.showPage()
                    c.save()
                    self.update_signal.emit(f"PDF '{pdf_filename}' gerado.")
            
            if self.generate_dxf:
                self.update_signal.emit("--- Gerando DXFs ---")
                output_folder = "DXF_Gerados"; os.makedirs(output_folder, exist_ok=True)
                zip_filename = f"{output_folder}/LOTE_DXF_{datetime.now():%Y%m%d_%H%M%S}.zip"
                with zipfile.ZipFile(zip_filename, 'w', zipfile.ZIP_DEFLATED) as zf:
                    for index, row in self.df.iterrows():
                        raw_data = row.to_dict()
                        prepared_data, error = dxf_engine.prepare_and_validate_dxf_data(raw_data)
                        if error: self.update_signal.emit(f"AVISO: Pulando DXF para '{raw_data.get('nome_arquivo')}': {error}"); continue
                        dxf_content, filename = dxf_engine.create_dxf_drawing(prepared_data)
                        if dxf_content: zf.writestr(filename, dxf_content)
                        else: self.update_signal.emit(f"ERRO: Falha ao gerar DXF para '{raw_data.get('nome_arquivo')}'.")
                        self.progress_signal.emit(int((index + 1) / len(self.df) * 100))
                self.update_signal.emit(f"Arquivo ZIP '{zip_filename}' gerado.")

            self.finished_signal.emit(True, "Processamento concluído com sucesso!")
        except Exception as e:
            self.finished_signal.emit(False, f"Erro crítico no processamento: {str(e)}")

# ==============================================
# INTERFACE GRÁFICA (MAIN WINDOW)
# ==============================================
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Gerador de Desenhos Técnicos e DXF")
        self.setGeometry(100, 100, 1100, 850)
        
        self.code_generator = CodeGenerator()
        self.colunas_df = ['nome_arquivo', 'forma', 'espessura', 'qtd', 'largura', 'altura', 'diametro', 'rt_base', 'rt_height', 'furos']
        self.manual_df = pd.DataFrame(columns=self.colunas_df)
        self.excel_df = pd.DataFrame(columns=self.colunas_df)
        self.furos_atuais = []

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        top_h_layout = QHBoxLayout()
        left_v_layout = QVBoxLayout()
        
        # Grupo 1: Arquivo
        file_group = QGroupBox("1. Carregar Planilha Excel (Opcional)")
        file_layout = QVBoxLayout()
        self.file_label = QLabel("Nenhuma planilha selecionada")
        file_button_layout = QHBoxLayout()
        self.select_file_btn = QPushButton("Selecionar Planilha")
        self.clear_excel_btn = QPushButton("Limpar Planilha")
        file_button_layout.addWidget(self.select_file_btn)
        file_button_layout.addWidget(self.clear_excel_btn)
        file_layout.addWidget(self.file_label)
        file_layout.addLayout(file_button_layout)
        file_group.setLayout(file_layout)
        left_v_layout.addWidget(file_group)

        # Grupo 2: Informações da Peça
        manual_group = QGroupBox("2. Informações da Peça")
        manual_layout = QFormLayout()
        self.projeto_input = QLineEdit()
        manual_layout.addRow("Nº do Projeto:", self.projeto_input)
        self.nome_input = QLineEdit()
        self.generate_code_btn = QPushButton("Gerar Código") # <<-- Mudei aqui para poder acessar depois
        name_layout = QHBoxLayout()
        name_layout.addWidget(self.nome_input)
        name_layout.addWidget(self.generate_code_btn)
        manual_layout.addRow("Nome/ID da Peça:", name_layout)
        self.forma_combo = QComboBox()
        self.forma_combo.addItems(['rectangle', 'circle', 'right_triangle'])
        self.espessura_input = QLineEdit()
        self.qtd_input = QLineEdit()
        self.largura_input, self.altura_input = QLineEdit(), QLineEdit()
        self.diametro_input, self.rt_base_input, self.rt_height_input = QLineEdit(), QLineEdit(), QLineEdit()
        manual_layout.addRow("Forma:", self.forma_combo)
        manual_layout.addRow("Espessura (mm):", self.espessura_input)
        manual_layout.addRow("Quantidade:", self.qtd_input)
        self.largura_row = [QLabel("Largura:"), self.largura_input]; manual_layout.addRow(self.largura_row[0], self.largura_row[1])
        self.altura_row = [QLabel("Altura:"), self.altura_input]; manual_layout.addRow(self.altura_row[0], self.altura_row[1])
        self.diametro_row = [QLabel("Diâmetro:"), self.diametro_input]; manual_layout.addRow(self.diametro_row[0], self.diametro_row[1])
        self.rt_base_row = [QLabel("Base Triângulo:"), self.rt_base_input]; manual_layout.addRow(self.rt_base_row[0], self.rt_base_row[1])
        self.rt_height_row = [QLabel("Altura Triângulo:"), self.rt_height_input]; manual_layout.addRow(self.rt_height_row[0], self.rt_height_row[1])
        manual_group.setLayout(manual_layout)
        left_v_layout.addWidget(manual_group)
        left_v_layout.addStretch()
        top_h_layout.addLayout(left_v_layout)

        # Grupo 3: Furos
        furos_main_group = QGroupBox("3. Adicionar Furos à Peça")
        furos_main_layout = QVBoxLayout()
        self.rep_group = QGroupBox("Furação Rápida (Replicar Furos nos Cantos)")
        rep_layout = QFormLayout()
        self.rep_diam_input, self.rep_offset_input = QLineEdit(), QLineEdit()
        rep_layout.addRow("Diâmetro dos Furos:", self.rep_diam_input)
        rep_layout.addRow("Distância da Borda (Offset):", self.rep_offset_input)
        self.replicate_btn = QPushButton("Replicar Furos")
        rep_layout.addRow(self.replicate_btn)
        self.rep_group.setLayout(rep_layout)
        furos_main_layout.addWidget(self.rep_group)
        man_group = QGroupBox("Furos Manuais")
        man_layout = QVBoxLayout()
        man_form_layout = QFormLayout()
        self.diametro_furo_input, self.pos_x_input, self.pos_y_input = QLineEdit(), QLineEdit(), QLineEdit()
        man_form_layout.addRow("Diâmetro do Furo:", self.diametro_furo_input)
        man_form_layout.addRow("Posição X:", self.pos_x_input)
        man_form_layout.addRow("Posição Y:", self.pos_y_input)
        self.add_furo_btn = QPushButton("Adicionar Furo Manual")
        man_layout.addLayout(man_form_layout)
        man_layout.addWidget(self.add_furo_btn)
        self.furos_table = QTableWidget(0, 4)
        self.furos_table.setHorizontalHeaderLabels(["Diâmetro", "Pos X", "Pos Y", "Ação"])
        man_layout.addWidget(self.furos_table)
        man_group.setLayout(man_layout)
        furos_main_layout.addWidget(man_group)
        furos_main_group.setLayout(furos_main_layout)
        top_h_layout.addWidget(furos_main_group, stretch=1)
        main_layout.addLayout(top_h_layout)
        self.add_piece_btn = QPushButton("Adicionar Peça à Lista de Produção")
        main_layout.addWidget(self.add_piece_btn)

        # Grupo 4: Lista de Produção e Ações
        list_group = QGroupBox("4. Lista de Peças para Produção")
        list_layout = QVBoxLayout()
        self.pieces_table = QTableWidget()
        self.table_headers = [col.replace('_', ' ').title() for col in self.colunas_df] + ["Ações"]
        self.pieces_table.setColumnCount(len(self.table_headers))
        self.pieces_table.setHorizontalHeaderLabels(self.table_headers)
        list_layout.addWidget(self.pieces_table)
        process_buttons_layout = QHBoxLayout()
        self.process_pdf_btn = QPushButton("Gerar PDFs")
        self.process_dxf_btn = QPushButton("Gerar DXFs")
        self.process_all_btn = QPushButton("Gerar PDFs e DXFs")
        process_buttons_layout.addWidget(self.process_pdf_btn)
        process_buttons_layout.addWidget(self.process_dxf_btn)
        process_buttons_layout.addWidget(self.process_all_btn)
        list_layout.addLayout(process_buttons_layout)
        list_group.setLayout(list_layout)
        main_layout.addWidget(list_group, stretch=1)
        
        # Status, Progresso e Log
        self.progress_bar = QProgressBar(); main_layout.addWidget(self.progress_bar)
        log_group = QGroupBox("Log de Execução")
        log_layout = QVBoxLayout()
        self.log_text = QTextEdit(); log_layout.addWidget(self.log_text)
        log_group.setLayout(log_layout)
        main_layout.addWidget(log_group)
        self.statusBar().showMessage("Pronto")
        
        # Conexões de Sinais e Slots
        self.select_file_btn.clicked.connect(self.select_file)
        self.clear_excel_btn.clicked.connect(self.clear_excel_data)
        self.generate_code_btn.clicked.connect(self.generate_piece_code)
        self.add_piece_btn.clicked.connect(self.add_manual_piece)
        self.forma_combo.currentTextChanged.connect(self.update_dimension_fields)
        self.replicate_btn.clicked.connect(self.replicate_holes)
        self.add_furo_btn.clicked.connect(self.add_furo_temp)
        self.process_pdf_btn.clicked.connect(self.start_pdf_generation)
        self.process_dxf_btn.clicked.connect(self.start_dxf_generation)
        self.process_all_btn.clicked.connect(self.start_all_generation)
        
        # --- CORREÇÃO: Define o estado inicial correto dos botões ---
        self.set_initial_button_state()
        self.update_dimension_fields(self.forma_combo.currentText())

    def set_initial_button_state(self):
        """Define o estado dos botões quando o aplicativo é iniciado."""
        # Botões que sempre começam habilitados
        self.select_file_btn.setEnabled(True)
        self.generate_code_btn.setEnabled(True)
        self.add_piece_btn.setEnabled(True)
        self.replicate_btn.setEnabled(True)
        self.add_furo_btn.setEnabled(True)

        # Botões que dependem de algum estado (começam desabilitados)
        self.clear_excel_btn.setEnabled(False)
        self.process_pdf_btn.setEnabled(False)
        self.process_dxf_btn.setEnabled(False)
        self.process_all_btn.setEnabled(False)
        self.progress_bar.setVisible(False)

    def set_buttons_enabled_on_process(self, enabled):
        """Controla o estado dos botões DURANTE o processamento."""
        self.select_file_btn.setEnabled(enabled)
        self.clear_excel_btn.setEnabled(enabled and not self.excel_df.empty)
        self.generate_code_btn.setEnabled(enabled)
        self.add_piece_btn.setEnabled(enabled)
        self.replicate_btn.setEnabled(enabled)
        self.add_furo_btn.setEnabled(enabled)
        
        has_items = not (self.excel_df.empty and self.manual_df.empty)
        self.process_pdf_btn.setEnabled(enabled and has_items)
        self.process_dxf_btn.setEnabled(enabled and has_items)
        self.process_all_btn.setEnabled(enabled and has_items)

    def start_pdf_generation(self): self.start_processing(generate_pdf=True, generate_dxf=False)
    def start_dxf_generation(self): self.start_processing(generate_pdf=False, generate_dxf=True)
    def start_all_generation(self): self.start_processing(generate_pdf=True, generate_dxf=True)

    def start_processing(self, generate_pdf, generate_dxf):
        combined_df = pd.concat([self.excel_df, self.manual_df], ignore_index=True)
        if combined_df.empty: QMessageBox.warning(self, "Aviso", "A lista de peças está vazia."); return
        
        self.set_buttons_enabled_on_process(False) # <<<-- Usa a nova função
        
        self.progress_bar.setVisible(True); self.progress_bar.setValue(0); self.log_text.clear()
        self.process_thread = ProcessThread(combined_df.copy(), generate_pdf, generate_dxf)
        self.process_thread.update_signal.connect(self.log_text.append)
        self.process_thread.progress_signal.connect(self.progress_bar.setValue)
        self.process_thread.finished_signal.connect(self.processing_finished)
        self.process_thread.start()

    def processing_finished(self, success, message):
        self.set_buttons_enabled_on_process(True) # <<<-- Usa a nova função
        self.progress_bar.setVisible(False)
        msgBox = QMessageBox.information if success else QMessageBox.critical
        msgBox(self, "Processamento Concluído" if success else "Erro", message)
        self.statusBar().showMessage("Pronto")
    
    def update_table_display(self):
        # Habilita botões de processo apenas se a lista combinada não estiver vazia
        has_items = not (self.excel_df.empty and self.manual_df.empty)
        self.process_pdf_btn.setEnabled(has_items)
        self.process_dxf_btn.setEnabled(has_items)
        self.process_all_btn.setEnabled(has_items)

        # Habilita o botão de limpar apenas se o DF do excel não estiver vazio
        self.clear_excel_btn.setEnabled(not self.excel_df.empty)

        combined_df = pd.concat([self.excel_df, self.manual_df], ignore_index=True)
        self.pieces_table.setRowCount(0)
        if combined_df.empty: return

        self.pieces_table.setRowCount(len(combined_df))
        for i, row in combined_df.iterrows():
            for j, col in enumerate(self.colunas_df):
                value = row[col]
                display_value = f"{len(value)} Furo(s)" if col == 'furos' and isinstance(value, list) else ('-' if pd.isna(value) or value == 0 else str(value))
                self.pieces_table.setItem(i, j, QTableWidgetItem(display_value))
            action_widget = QWidget(); action_layout = QHBoxLayout(action_widget); action_layout.setContentsMargins(5, 0, 5, 0)
            edit_btn = QPushButton("Editar"); delete_btn = QPushButton("Excluir")
            edit_btn.clicked.connect(lambda _, r=i: self.edit_row(r))
            delete_btn.clicked.connect(lambda _, r=i: self.delete_row(r))
            action_layout.addWidget(edit_btn); action_layout.addWidget(delete_btn)
            self.pieces_table.setCellWidget(i, len(self.colunas_df), action_widget)
        self.pieces_table.resizeColumnsToContents()
    
    # ... (O resto das suas funções permanece o mesmo) ...
    def edit_row(self, row_index):
        len_excel = len(self.excel_df)
        df_source = self.excel_df if row_index < len_excel else self.manual_df
        local_index = row_index if row_index < len_excel else row_index - len_excel
        piece_data = df_source.iloc[local_index]
        self.nome_input.setText(str(piece_data.get('nome_arquivo', ''))); self.espessura_input.setText(str(piece_data.get('espessura', ''))); self.qtd_input.setText(str(piece_data.get('qtd', '')))
        shape = piece_data.get('forma', 'rectangle'); index = self.forma_combo.findText(shape, Qt.MatchFixedString)
        if index >= 0: self.forma_combo.setCurrentIndex(index)
        self.largura_input.setText(str(piece_data.get('largura', ''))); self.altura_input.setText(str(piece_data.get('altura', ''))); self.diametro_input.setText(str(piece_data.get('diametro', '')))
        self.rt_base_input.setText(str(piece_data.get('rt_base', ''))); self.rt_height_input.setText(str(piece_data.get('rt_height', '')))
        self.furos_atuais = piece_data.get('furos', []).copy() if isinstance(piece_data.get('furos'), list) else []
        self.update_furos_table()
        df_source.drop(df_source.index[local_index], inplace=True); df_source.reset_index(drop=True, inplace=True)
        self.log_text.append(f"Peça '{piece_data['nome_arquivo']}' carregada para edição."); self.update_table_display()

    def delete_row(self, row_index):
        len_excel = len(self.excel_df)
        df_source = self.excel_df if row_index < len_excel else self.manual_df
        local_index = row_index if row_index < len_excel else row_index - len_excel
        piece_name = df_source.iloc[local_index]['nome_arquivo']
        df_source.drop(df_source.index[local_index], inplace=True); df_source.reset_index(drop=True, inplace=True)
        self.log_text.append(f"Peça '{piece_name}' removida."); self.update_table_display()

    def generate_piece_code(self):
        project_number = self.projeto_input.text().strip()
        if not project_number: QMessageBox.warning(self, "Campo Obrigatório", "Insira o 'Nº do Projeto' para gerar um código."); return
        new_code = self.code_generator.generate_new_code(project_number, prefix='DES')
        if new_code: self.nome_input.setText(new_code); self.log_text.append(f"Código '{new_code}' gerado para o projeto '{project_number}'.")

    def add_manual_piece(self):
        try:
            nome = self.nome_input.text().strip()
            if not nome: QMessageBox.warning(self, "Campo Obrigatório", "'Nome/ID da Peça' é obrigatório."); return
            new_piece = {'furos': self.furos_atuais.copy()}
            for col in self.colunas_df:
                if col != 'furos': new_piece[col] = 0
            new_piece.update({'nome_arquivo': nome, 'forma': self.forma_combo.currentText()})
            for field, key in [(self.espessura_input, 'espessura'), (self.qtd_input, 'qtd'), (self.largura_input, 'largura'), (self.altura_input, 'altura'), (self.diametro_input, 'diametro'), (self.rt_base_input, 'rt_base'), (self.rt_height_input, 'rt_height')]:
                new_piece[key] = float(field.text().replace(',', '.')) if field.text() else 0
            self.manual_df = pd.concat([self.manual_df, pd.DataFrame([new_piece])], ignore_index=True)
            self.log_text.append(f"Peça '{nome}' adicionada/atualizada na lista.")
            for field in [self.nome_input, self.espessura_input, self.qtd_input, self.largura_input, self.altura_input, self.diametro_input, self.rt_base_input, self.rt_height_input, self.projeto_input]: field.clear()
            self.furos_atuais = []; self.update_furos_table(); self.update_table_display()
        except ValueError: QMessageBox.critical(self, "Erro de Valor", "Campos numéricos devem conter números válidos.")

    def select_file(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Selecionar Planilha", "", "Excel Files (*.xlsx *.xls)")
        if file_path:
            try:
                df = pd.read_excel(file_path, header=0, decimal=','); df.columns = df.columns.str.strip().str.lower()
                df = df.loc[:, ~df.columns.duplicated()]
                for col in self.colunas_df:
                    if col not in df.columns: df[col] = pd.NA
                if 'furos' in df.columns: df['furos'] = df['furos'].apply(lambda x: x if isinstance(x, list) else [])
                self.excel_df = df[self.colunas_df]
                self.file_label.setText(f"Planilha: {os.path.basename(file_path)}"); self.update_table_display()
            except Exception as e: QMessageBox.critical(self, "Erro de Leitura", f"Falha ao ler o arquivo: {e}")

    def clear_excel_data(self):
        self.excel_df = pd.DataFrame(columns=self.colunas_df); self.file_label.setText("Nenhuma planilha selecionada"); self.update_table_display()
    
    def replicate_holes(self):
        try:
            if self.forma_combo.currentText() != 'rectangle': QMessageBox.warning(self, "Função Indisponível", "Replicação de furos disponível apenas para Retângulos."); return
            largura, altura = float(self.largura_input.text().replace(',', '.')), float(self.altura_input.text().replace(',', '.'))
            diam, offset = float(self.rep_diam_input.text().replace(',', '.')), float(self.rep_offset_input.text().replace(',', '.'))
            if (offset * 2) >= largura or (offset * 2) >= altura: QMessageBox.warning(self, "Offset Inválido", "O dobro do offset excede as dimensões da peça."); return
            furos_replicados = [{'diam': diam, 'x': offset, 'y': offset}, {'diam': diam, 'x': largura - offset, 'y': offset}, {'diam': diam, 'x': largura - offset, 'y': altura - offset}, {'diam': diam, 'x': offset, 'y': altura - offset}]
            self.furos_atuais.extend(furos_replicados); self.update_furos_table()
        except ValueError: QMessageBox.critical(self, "Erro de Valor", "Largura, Altura, Diâmetro e Offset devem ser números válidos.")

    def update_dimension_fields(self, shape):
        shape = shape.lower()
        is_rect = shape == 'rectangle'; is_circ = shape == 'circle'; is_tri = shape == 'right_triangle'
        for w in self.largura_row + self.altura_row: w.setVisible(is_rect)
        for w in self.diametro_row: w.setVisible(is_circ)
        for w in self.rt_base_row + self.rt_height_row: w.setVisible(is_tri)
        self.rep_group.setEnabled(is_rect)

    def add_furo_temp(self):
        try:
            diam, pos_x, pos_y = float(self.diametro_furo_input.text().replace(',', '.')), float(self.pos_x_input.text().replace(',', '.')), float(self.pos_y_input.text().replace(',', '.'))
            if diam <= 0: QMessageBox.warning(self, "Valor Inválido", "Diâmetro do furo deve ser maior que zero."); return
            self.furos_atuais.append({'diam': diam, 'x': pos_x, 'y': pos_y}); self.update_furos_table()
            for field in [self.diametro_furo_input, self.pos_x_input, self.pos_y_input]: field.clear()
        except ValueError: QMessageBox.critical(self, "Erro de Valor", "Campos de furo devem ser números válidos.")

    def update_furos_table(self):
        self.furos_table.setRowCount(0); self.furos_table.setRowCount(len(self.furos_atuais))
        for i, furo in enumerate(self.furos_atuais):
            self.furos_table.setItem(i, 0, QTableWidgetItem(str(furo['diam']))); self.furos_table.setItem(i, 1, QTableWidgetItem(str(furo['x']))); self.furos_table.setItem(i, 2, QTableWidgetItem(str(furo['y'])))
            delete_btn = QPushButton("Excluir"); delete_btn.clicked.connect(lambda _, r=i: self.delete_furo_temp(r)); self.furos_table.setCellWidget(i, 3, delete_btn)
        self.furos_table.resizeColumnsToContents()

    def delete_furo_temp(self, row_index):
        del self.furos_atuais[row_index]; self.update_furos_table()

# ==============================================
# PONTO DE ENTRADA DA APLICAÇÃO
# ==============================================
def main():
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()